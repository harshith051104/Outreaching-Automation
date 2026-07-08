"""
CSV parsing utilities for lead imports.

Supports standard CSV formats with flexible column name matching.
"""

import csv
import io
import sys
from typing import Any

from app.config.constants import MAX_CSV_ROWS

# Increase python's csv field limit to prevent crash on large cells/descriptions
csv.field_size_limit(13107200)


COLUMN_ALIASES: dict[str, str] = {
    "email": "email",
    "emailaddress": "email",
    "email_address": "email",
    "firstname": "first_name",
    "first_name": "first_name",
    "lastname": "last_name",
    "last_name": "last_name",
    "name": "first_name",
    "fullname": "first_name",
    "investor_name": "first_name",
    "investors_name": "first_name",
    "company": "company",
    "company_name": "company",
    "organization": "company",
    "title": "title",
    "job_title": "title",
    "role": "title",
    "phone": "phone",
    "phone_number": "phone",
    "linkedin": "linkedin_url",
    "linkedin_url": "linkedin_url",
    "linkedin_profile": "linkedin_url",
    "website": "website",
    "url": "website",
    "industry": "industry",
    "focus": "focus",
    "investor_focus": "focus",
    "investorfocus": "focus",
    "investor focus": "focus"
}


def normalize_column_name(col: str) -> str:
    """Map a CSV column header to the canonical key name with robust similarity matching."""
    if not col:
        return ""
    
    import re
    # Replace spaces, hyphens, slashes, and backslashes with underscores
    col_lower = str(col).lower().strip()
    col_replaced = re.sub(r'[\s\-/\\]+', '_', col_lower)
    # Remove all other non-alphanumeric/non-underscore characters (e.g. parentheses, punctuation, dollar signs)
    col_clean = re.sub(r'[^a-z0-9_]', '', col_replaced)
    # Collapse multiple consecutive underscores and strip trailing/leading underscores
    col_clean = re.sub(r'_+', '_', col_clean).strip('_')

    # 1. Check direct aliases first
    canonical = COLUMN_ALIASES.get(col_clean)
    if canonical:
        return canonical
    
    # 2. Check similarity mappings for standard fields and custom fields
    # Email similarity - exact match only to avoid false matches on "email_status" etc.
    if col_clean in ["email", "emailaddress", "email_address", "mail", "e_mail", "e-mail", "e_mail_address"]:
        return "email"
        
    # First name similarity
    if col_clean in ["first_name", "firstname", "given_name"]:
        return "first_name"
        
    # Last name similarity
    if col_clean in ["last_name", "lastname", "family_name"]:
        return "last_name"
        
    # Name similarity
    if col_clean in ["name", "full_name", "fullname", "contact_name", "person_name"]:
        return "first_name"
        
    # Company similarity
    if "company" in col_clean or col_clean in ["organization", "org"]:
        return "company"
        
    # Role/Title similarity
    if col_clean in ["role", "job_title", "title", "position", "headline"]:
        return "title"
        
    # Website similarity
    if any(x in col_clean for x in ["website", "site_url", "web_url", "link_to_website"]) or col_clean in ["url", "site"]:
        return "website"
        
    # LinkedIn similarity — exact matches only to avoid "LinkedIn Status" → linkedin_url
    if col_clean in ["linkedin", "linkedin_url", "linkedin_profile", "linkedin_profile_url", "profile_url", "profile", "linkedin_link", "linkedin_url_link"]:
        return "linkedin_url"
        
    # Focus similarity
    if "focus" in col_clean:
        return "focus"
        
    # Avg check size
    if any(x in col_clean for x in ["avg_check", "average_check", "check_size"]):
        return "avg_check_size"
    
    # Total Under Management
    if any(x in col_clean for x in ["under_management", "aum", "assets_under"]):
        return "total_under_management"
        
    # Firm Description
    if any(x in col_clean for x in ["firm_desc", "company_desc", "business_desc", "firm_description", "company_description"]):
        return "firm_description"
        
    # Portfolio Companies
    if any(x in col_clean for x in ["portfolio_companies", "portfolio_company", "portcos", "portfolio"]):
        return "portfolio_companies"
        
    # Personal Paragraph
    if any(x in col_clean for x in ["personal_paragraph", "personal_note", "personalization"]):
        return "personal_paragraph"
        
    return col_clean


def parse_leads_csv(file_content: bytes) -> list[dict[str, Any]]:
    """
    Parse a CSV file containing lead data.

    Args:
        file_content: Raw bytes of the CSV file.

    Returns:
        List of dictionaries, each representing a lead row.

    Raises:
        ValueError: If the CSV is malformed or exceeds MAX_CSV_ROWS.
    """
    try:
        text = file_content.decode("utf-8")
    except UnicodeDecodeError:
        for encoding in ["latin-1", "cp1252"]:
            try:
                text = file_content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            raise ValueError("Unable to decode CSV file. Ensure UTF-8 encoding.")

    reader = csv.DictReader(io.StringIO(text))

    if not reader.fieldnames:
        raise ValueError("CSV file is empty or has no headers.")

    canonical_fields = [normalize_column_name(f) for f in reader.fieldnames]
    reader.fieldnames = canonical_fields

    rows: list[dict[str, Any]] = []
    for i, row in enumerate(reader, start=1):
        if i > MAX_CSV_ROWS:
            break

        if not (row.get("email") or "").strip():
            continue

        cleaned = {k: (v.strip() if isinstance(v, str) else "") for k, v in row.items() if k is not None}
        rows.append(cleaned)

    return rows


def parse_leads_xlsx(file_content: bytes) -> list[dict[str, Any]]:
    # ponytail: Excel XLSX parsing helper
    try:
        import openpyxl
    except ImportError:
        raise ValueError("Excel parsing library (openpyxl) is not installed.")
    import io
    wb = openpyxl.load_workbook(filename=io.BytesIO(file_content), data_only=True, read_only=True)
    if not wb.sheetnames:
        raise ValueError("Excel file is empty.")
    sheet = wb.active or wb[wb.sheetnames[0]]
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        headers = next(rows_iter)
    except StopIteration:
        raise ValueError("Excel sheet is empty.")
    if not headers:
        raise ValueError("Excel sheet has no headers.")
    headers = [str(h).strip() for h in headers if h is not None]
    canonical_fields = [normalize_column_name(h) for h in headers]
    rows = []
    for i, row in enumerate(rows_iter, start=1):
        if i > MAX_CSV_ROWS:
            break
        row_dict = {}
        for idx, field in enumerate(canonical_fields):
            val = row[idx] if idx < len(row) else None
            row_dict[field] = str(val).strip() if val is not None else ""
        if not row_dict.get("email"):
            continue
        rows.append(row_dict)
    wb.close()
    return rows


async def parse_leads_pdf(file_content: bytes) -> list[dict[str, Any]]:
    # ponytail: PDF parser extracting raw text and structuring via LLM
    try:
        import pypdf
    except ImportError:
        raise ValueError("PDF parsing library (pypdf) is not installed.")
    import io
    import json
    from app.config.groq_config import get_groq_client
    from app.config.settings import settings
    import logging
    logger = logging.getLogger(__name__)
    try:
        pdf = pypdf.PdfReader(io.BytesIO(file_content))
        text_content = []
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                text_content.append(text)
        raw_text = "\n".join(text_content).strip()
    except Exception as exc:
        raise ValueError(f"Failed to read PDF: {exc}")
    if not raw_text:
        raise ValueError("No text could be extracted from PDF.")
    system_prompt = (
        "You are an expert data extraction assistant. "
        "Extract all lead contacts from the provided unstructured text. "
        "For each lead, extract: name, email, company, title, phone, linkedin_url, website, focus. "
        "Match column headings to these keys where possible. "
        "Return ONLY a valid JSON list of objects. "
        "Do not include any chat formatting, notes, or codeblocks outside the JSON array."
    )
    user_prompt = f"Extract leads from the following text:\n\n{raw_text}"
    try:
        client = get_groq_client(None)
        model = getattr(settings, "GROQ_MODEL", "llama-3.3-70b-versatile")
        import asyncio
        resp = await asyncio.to_thread(
            client.chat.completions.create,
            model=model,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt}
            ],
            temperature=0.1,
            max_tokens=2540
        )
        content = (resp.choices[0].message.content or "").strip()
        if content.startswith("```"):
            lines = content.splitlines()
            if lines[0].startswith("```"):
                lines = lines[1:]
            if lines and lines[-1].startswith("```"):
                lines = lines[:-1]
            content = "\n".join(lines).strip()
        data = json.loads(content)
        if isinstance(data, list):
            normalized_rows = []
            for item in data:
                normalized_item = {}
                for k, v in item.items():
                    norm_k = normalize_column_name(k)
                    normalized_item[norm_k] = str(v).strip() if v is not None else ""
                normalized_rows.append(normalized_item)
            return normalized_rows
        return []
    except Exception as exc:
        logger.error("LLM lead extraction failed: %s", exc)
        raise ValueError(f"Failed to parse PDF text via AI: {exc}")


async def parse_leads_file(file_content: bytes, filename: str) -> list[dict[str, Any]]:
    # ponytail: central dynamic parser dispatch
    fn_lower = filename.lower()
    if fn_lower.endswith((".xlsx", ".xls")):
        import asyncio
        return await asyncio.to_thread(parse_leads_xlsx, file_content)
    elif fn_lower.endswith(".pdf"):
        return await parse_leads_pdf(file_content)
    else:
        return parse_leads_csv(file_content)


def validate_lead_row(row: dict[str, Any]) -> tuple[bool, str]:
    """
    Validate a single lead row.

    Returns:
        Tuple of (is_valid, error_message).
    """
    email = (row.get("email") or "").strip()
    if not email:
        return False, "Missing email field"

    if "@" not in email or "." not in email.split("@")[-1]:
        return False, f"Invalid email format: {email}"

    return True, ""